# handler.py

import json
import base64
from io import BytesIO
from openpyxl import load_workbook


def main(event, context):
    # Openpyxl

    file_content = event["content"]
    decrypted = base64.b64decode(file_content)
    wb = load_workbook(filename=BytesIO(decrypted))
    sheet = wb["questions"]
    rows = sheet.max_row

    message = 'Number of rows: {}!'.format(rows)

    body = {
        "message": message
    }
    response = {
        "statusCode": 200,
        "body": json.dumps(body)
    }
    return response


# This is for local testing the code without lambda
if __name__ == "__main__":
    main('', '')

